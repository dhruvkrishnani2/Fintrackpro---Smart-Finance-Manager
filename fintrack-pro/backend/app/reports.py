"""Excel / PDF report builders for FinTrack Pro.

Every function here returns an in-memory io.BytesIO buffer, ready to be
streamed back by a FastAPI route via StreamingResponse. Keeping the
rendering logic separate from the router keeps the endpoints thin and
makes the builders easy to unit test in isolation.
"""
import io
from datetime import datetime
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ---------------------------------------------------------------------------
# Shared palette (mirrors the frontend's ink / emerald / coral / gold theme)
# ---------------------------------------------------------------------------
INK = colors.HexColor("#1F2A24")
EMERALD = colors.HexColor("#16794F")
CORAL = colors.HexColor("#C0524B")
GOLD = colors.HexColor("#B8892E")
PAPER = colors.HexColor("#F7F4EE")
GREY_LINE = colors.HexColor("#DEDACF")

HEADER_FILL = "1F2A24"
STRIPE_FILL = "F7F4EE"
CURRENCY_FMT = "#,##0.00"


def _category_map(categories: Iterable) -> Dict[Optional[str], str]:
    return {c.id: c.name for c in categories}


def _fmt_money(value: float) -> str:
    return f"Rs {value:,.2f}"


def _fmt_date(value: Optional[datetime]) -> str:
    return value.strftime("%d %b %Y") if value else "-"


# ---------------------------------------------------------------------------
# Excel: transaction ledger export
# ---------------------------------------------------------------------------
def build_transactions_excel(transactions: List, categories: Iterable, title: str = "Transactions") -> io.BytesIO:
    cat_map = _category_map(categories)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    headers = ["Date", "Type", "Category", "Description", "Source", "Amount (signed)", "Recurring"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_income = 0.0
    total_expense = 0.0
    row_idx = 2
    for tx in transactions:
        is_income = tx.type.value == "income"
        signed_amount = tx.amount if is_income else -tx.amount
        if is_income:
            total_income += tx.amount
        else:
            total_expense += tx.amount

        ws.append([
            _fmt_date(tx.date),
            tx.type.value.capitalize(),
            cat_map.get(tx.category_id, "Uncategorized"),
            tx.description or "",
            tx.source or "",
            signed_amount,
            "Yes" if tx.is_recurring else "No",
        ])
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=STRIPE_FILL, end_color=STRIPE_FILL, fill_type="solid"
                )
        amount_cell = ws.cell(row=row_idx, column=6)
        amount_cell.number_format = CURRENCY_FMT
        amount_cell.font = Font(color="16794F" if is_income else "C0524B")
        row_idx += 1

    # Totals footer
    ws.append([])
    ws.append(["", "", "", "", "Total income", total_income, ""])
    ws.append(["", "", "", "", "Total expenses", -total_expense, ""])
    ws.append(["", "", "", "", "Net total", total_income - total_expense, ""])
    for r in range(row_idx + 1, row_idx + 4):
        ws.cell(row=r, column=5).font = Font(bold=True)
        c = ws.cell(row=r, column=6)
        c.font = Font(bold=True)
        c.number_format = CURRENCY_FMT

    widths = [14, 12, 20, 32, 16, 18, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF: transaction ledger export
# ---------------------------------------------------------------------------
def build_transactions_pdf(transactions: List, categories: Iterable, user_name: str, subtitle: str = "") -> io.BytesIO:
    cat_map = _category_map(categories)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], textColor=INK, fontSize=20, spaceAfter=2)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], textColor=colors.grey, fontSize=9, spaceAfter=14)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=18 * mm, bottomMargin=16 * mm, leftMargin=16 * mm, rightMargin=16 * mm,
    )

    story = [
        Paragraph("FinTrack Pro", ParagraphStyle("Brand", parent=styles["Normal"], textColor=EMERALD, fontSize=11, spaceAfter=4)),
        Paragraph("Transaction Statement", title_style),
        Paragraph(
            f"{user_name} &bull; Generated {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}"
            + (f" &bull; {subtitle}" if subtitle else ""),
            sub_style,
        ),
    ]

    table_data = [["Date", "Type", "Category", "Description", "Amount"]]
    total_income = 0.0
    total_expense = 0.0
    row_styles = []
    for i, tx in enumerate(transactions, start=1):
        is_income = tx.type.value == "income"
        if is_income:
            total_income += tx.amount
        else:
            total_expense += tx.amount
        amount_text = f"{'+' if is_income else '-'}{tx.amount:,.2f}"
        table_data.append([
            _fmt_date(tx.date),
            tx.type.value.capitalize(),
            cat_map.get(tx.category_id, "Uncategorized"),
            (tx.description or "-")[:40],
            amount_text,
        ])
        row_styles.append(("TEXTCOLOR", (4, i), (4, i), EMERALD if is_income else CORAL))

    if len(table_data) == 1:
        table_data.append(["-", "-", "-", "No transactions in range", "-"])

    tbl = Table(table_data, colWidths=[24 * mm, 20 * mm, 30 * mm, 62 * mm, 28 * mm], repeatRows=1)
    base_style = [
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PAPER]),
        ("LINEBELOW", (0, 0), (-1, 0), 0.75, INK),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.25, GREY_LINE),
    ] + row_styles
    tbl.setStyle(TableStyle(base_style))
    story.append(tbl)
    story.append(Spacer(1, 10 * mm))

    summary_data = [
        ["Total income", _fmt_money(total_income)],
        ["Total expenses", _fmt_money(total_expense)],
        ["Net total", _fmt_money(total_income - total_expense)],
    ]
    summary_tbl = Table(summary_data, colWidths=[40 * mm, 40 * mm], hAlign="RIGHT")
    summary_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (0, 2), (-1, 2), 0.75, INK),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(summary_tbl)

    doc.build(story)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Excel: full financial report (multi-sheet workbook)
# ---------------------------------------------------------------------------
def build_full_report_excel(
    user_name: str,
    summary: dict,
    breakdown: List[dict],
    trend: List[dict],
    transactions: List,
    budgets: List[dict],
    goals: List,
    categories: Iterable,
) -> io.BytesIO:
    cat_map = _category_map(categories)
    wb = Workbook()

    def style_header(ws, row=1, n_cols=2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "FinTrack Pro — Financial Report"
    ws["A1"].font = Font(bold=True, size=14, color=HEADER_FILL)
    ws["A2"] = f"{user_name} • Generated {datetime.utcnow().strftime('%d %b %Y %H:%M UTC')}"
    ws["A2"].font = Font(italic=True, color="666666")

    ws.append([])
    ws.append(["Metric", "Value"])
    style_header(ws, row=ws.max_row)
    for label, key in [
        ("Total balance", "total_balance"),
        ("Monthly income", "monthly_income"),
        ("Monthly expenses", "monthly_expenses"),
        ("Monthly savings", "monthly_savings"),
        ("Net cash flow", "net_cash_flow"),
    ]:
        ws.append([label, summary.get(key, 0)])
        ws.cell(row=ws.max_row, column=2).number_format = CURRENCY_FMT
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20

    # ---- Category breakdown sheet ----
    ws2 = wb.create_sheet("Category Breakdown")
    ws2.append(["Category", "Total spent"])
    style_header(ws2, n_cols=2)
    for row in breakdown:
        ws2.append([row["category"], row["total"]])
        ws2.cell(row=ws2.max_row, column=2).number_format = CURRENCY_FMT
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 18

    # ---- Trend sheet ----
    ws3 = wb.create_sheet("Monthly Trend")
    ws3.append(["Month", "Income", "Expenses", "Net"])
    style_header(ws3, n_cols=4)
    for row in trend:
        net = row["income"] - row["expenses"]
        ws3.append([row["label"], row["income"], row["expenses"], net])
        for col in (2, 3, 4):
            ws3.cell(row=ws3.max_row, column=col).number_format = CURRENCY_FMT
    for col, w in zip("ABCD", [14, 16, 16, 16]):
        ws3.column_dimensions[col].width = w

    # ---- Budgets sheet ----
    ws4 = wb.create_sheet("Budgets")
    ws4.append(["Category", "Month", "Year", "Limit", "Spent", "Remaining"])
    style_header(ws4, n_cols=6)
    for b in budgets:
        remaining = b["limit_amount"] - b["spent"]
        ws4.append([
            cat_map.get(b["category_id"], "Uncategorized"),
            b["month"], b["year"], b["limit_amount"], b["spent"], remaining,
        ])
        for col in (4, 5, 6):
            ws4.cell(row=ws4.max_row, column=col).number_format = CURRENCY_FMT
        if remaining < 0:
            ws4.cell(row=ws4.max_row, column=6).font = Font(color="C0524B", bold=True)
    for col, w in zip("ABCDEF", [20, 8, 8, 14, 14, 14]):
        ws4.column_dimensions[col].width = w

    # ---- Goals sheet ----
    ws5 = wb.create_sheet("Goals")
    ws5.append(["Goal", "Target", "Current", "Progress %", "Target date", "Status"])
    style_header(ws5, n_cols=6)
    for g in goals:
        progress = (g.current_amount / g.target_amount * 100) if g.target_amount else 0
        ws5.append([
            g.name, g.target_amount, g.current_amount, round(progress, 1),
            _fmt_date(g.target_date), g.status.value.capitalize(),
        ])
        ws5.cell(row=ws5.max_row, column=2).number_format = CURRENCY_FMT
        ws5.cell(row=ws5.max_row, column=3).number_format = CURRENCY_FMT
    for col, w in zip("ABCDEF", [24, 14, 14, 12, 14, 12]):
        ws5.column_dimensions[col].width = w

    # ---- Transactions sheet ----
    ws6 = wb.create_sheet("Transactions")
    ws6.append(["Date", "Type", "Category", "Description", "Source", "Amount (signed)"])
    style_header(ws6, n_cols=6)
    for tx in transactions:
        is_income = tx.type.value == "income"
        signed = tx.amount if is_income else -tx.amount
        ws6.append([
            _fmt_date(tx.date), tx.type.value.capitalize(),
            cat_map.get(tx.category_id, "Uncategorized"),
            tx.description or "", tx.source or "", signed,
        ])
        cell = ws6.cell(row=ws6.max_row, column=6)
        cell.number_format = CURRENCY_FMT
        cell.font = Font(color="16794F" if is_income else "C0524B")
    for col, w in zip("ABCDEF", [14, 12, 20, 32, 16, 16]):
        ws6.column_dimensions[col].width = w
    ws6.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF: full financial report
# ---------------------------------------------------------------------------
def build_full_report_pdf(
    user_name: str,
    summary: dict,
    breakdown: List[dict],
    trend: List[dict],
    budgets: List[dict],
    goals: List,
    categories: Iterable,
) -> io.BytesIO:
    cat_map = _category_map(categories)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], textColor=INK, fontSize=20, spaceAfter=2)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], textColor=colors.grey, fontSize=9, spaceAfter=14)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"], textColor=INK, fontSize=13, spaceBefore=14, spaceAfter=6)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=18 * mm, bottomMargin=16 * mm, leftMargin=16 * mm, rightMargin=16 * mm,
    )

    story = [
        Paragraph("FinTrack Pro", ParagraphStyle("Brand", parent=styles["Normal"], textColor=EMERALD, fontSize=11, spaceAfter=4)),
        Paragraph("Financial Report", title_style),
        Paragraph(f"{user_name} &bull; Generated {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}", sub_style),
    ]

    # Summary cards (as a table)
    summary_rows = [
        ["Total balance", "Monthly income", "Monthly expenses", "Monthly savings"],
        [
            _fmt_money(summary.get("total_balance", 0)),
            _fmt_money(summary.get("monthly_income", 0)),
            _fmt_money(summary.get("monthly_expenses", 0)),
            _fmt_money(summary.get("monthly_savings", 0)),
        ],
    ]
    sum_tbl = Table(summary_rows, colWidths=[42 * mm] * 4)
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PAPER),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.grey),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, 1), 12),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 1), (0, 1), INK),
        ("TEXTCOLOR", (1, 1), (1, 1), EMERALD),
        ("TEXTCOLOR", (2, 1), (2, 1), CORAL),
        ("TEXTCOLOR", (3, 1), (3, 1), GOLD),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.25, GREY_LINE),
        ("BOX", (0, 0), (-1, -1), 0.5, GREY_LINE),
    ]))
    story.append(sum_tbl)

    # Category breakdown
    story.append(Paragraph("Spending by category", h2_style))
    if breakdown:
        cat_data = [["Category", "Total"]] + [[b["category"], _fmt_money(b["total"])] for b in breakdown]
        cat_tbl = Table(cat_data, colWidths=[100 * mm, 40 * mm])
        cat_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), INK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PAPER]),
            ("GRID", (0, 0), (-1, -1), 0.25, GREY_LINE),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(cat_tbl)
    else:
        story.append(Paragraph("No expense data yet.", styles["Normal"]))

    # Monthly trend
    story.append(Paragraph("Income vs. expenses trend", h2_style))
    if trend:
        trend_data = [["Month", "Income", "Expenses", "Net"]] + [
            [t["label"], _fmt_money(t["income"]), _fmt_money(t["expenses"]), _fmt_money(t["income"] - t["expenses"])]
            for t in trend
        ]
        trend_tbl = Table(trend_data, colWidths=[35 * mm] * 4)
        trend_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), INK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PAPER]),
            ("GRID", (0, 0), (-1, -1), 0.25, GREY_LINE),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(trend_tbl)
    else:
        story.append(Paragraph("Not enough history yet.", styles["Normal"]))

    # Budgets
    story.append(Paragraph("Budgets this month", h2_style))
    if budgets:
        b_data = [["Category", "Limit", "Spent", "Remaining"]]
        b_styles = []
        for i, b in enumerate(budgets, start=1):
            remaining = b["limit_amount"] - b["spent"]
            b_data.append([
                cat_map.get(b["category_id"], "Uncategorized"),
                _fmt_money(b["limit_amount"]), _fmt_money(b["spent"]), _fmt_money(remaining),
            ])
            b_styles.append(("TEXTCOLOR", (3, i), (3, i), CORAL if remaining < 0 else EMERALD))
        b_tbl = Table(b_data, colWidths=[50 * mm] * 4)
        b_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), INK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PAPER]),
            ("GRID", (0, 0), (-1, -1), 0.25, GREY_LINE),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ] + b_styles))
        story.append(b_tbl)
    else:
        story.append(Paragraph("No budgets set for this month.", styles["Normal"]))

    # Goals
    story.append(Paragraph("Savings goals", h2_style))
    if goals:
        g_data = [["Goal", "Target", "Current", "Progress", "Status"]]
        for g in goals:
            progress = (g.current_amount / g.target_amount * 100) if g.target_amount else 0
            g_data.append([
                g.name, _fmt_money(g.target_amount), _fmt_money(g.current_amount),
                f"{progress:.0f}%", g.status.value.capitalize(),
            ])
        g_tbl = Table(g_data, colWidths=[45 * mm, 32 * mm, 32 * mm, 22 * mm, 22 * mm])
        g_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), INK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PAPER]),
            ("GRID", (0, 0), (-1, -1), 0.25, GREY_LINE),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(g_tbl)
    else:
        story.append(Paragraph("No savings goals yet.", styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    return buf
